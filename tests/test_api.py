"""Tests for the UniPulse AI FastAPI routes (``app.main``).

Covers the three registered endpoints:

- ``POST /api/v1/enrich/single``  — JSON payload and multipart PDF uploads.
- ``POST /api/v1/enrich/batch``   — CSV and Excel uploads (incl. failure rows).
- ``POST /api/v1/export/excel``   — workbook generation and error handling.

All network/LLM dependencies are faked via monkeypatched module-level
pipeline objects, so the suite runs fully offline and without API keys.
"""

from __future__ import annotations

import csv
import io
import json

import fitz
import openpyxl
import pytest
from fastapi.testclient import TestClient

from pathlib import Path

import app.main as api
from app.db.checkpoint_store import CheckpointStore
from app.schemas.enrichment import IndustrialAttribute
from app.services.rate_limiter import AdaptiveRateLimiter


@pytest.fixture(autouse=True)
def _reset_global_test_state(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    """Give every test a fresh global TPM limiter and isolated CheckpointStore."""
    monkeypatch.setattr(api, "_RATE_LIMITER", AdaptiveRateLimiter(max_rpm=120))
    test_db = str(tmp_path / "test_checkpoint.db")
    monkeypatch.setattr(api, "_checkpoint_store", CheckpointStore(db_path=test_db))

# ---------------------------------------------------------------------------
# Test doubles for the pipeline services
# ---------------------------------------------------------------------------


class FakeScraper:
    """Stand-in for ``WhitelistedSearchScraper`` with canned scrape results."""

    def __init__(self, results: list[dict[str, str]] | None = None) -> None:
        self.results = results or []

    async def search_and_scrape_async(self, manufacturer: str, part_number: str):
        return list(self.results)


class FakeExtractor:
    """Stand-in for ``StructuredExtractor`` that echoes deterministic specs."""

    def __init__(self, attributes: list[IndustrialAttribute] | None = None) -> None:
        self.attributes = attributes or [
            IndustrialAttribute(
                field_name="voltage",
                raw_value="120 VAC",
                normalized_value="120",
                unit="V",
                confidence_score=0.95,
                source_url="https://example.com/spec",
            ),
            IndustrialAttribute(
                field_name="airflow_cfm",
                raw_value="800 CFM",
                normalized_value="800",
                unit="CFM",
                confidence_score=0.9,
                source_url="https://example.com/spec",
            ),
        ]
        self.last_raw_text: str | None = None
        self.last_source_url: str | None = None
        self.last_category: str | None = None

    def extract_product_specs(self, raw_text: str, source_url: str, category: str):
        self.last_raw_text = raw_text
        self.last_source_url = source_url
        self.last_category = category
        return [
            attribute.model_copy(update={"source_url": source_url})
            for attribute in self.attributes
        ]


_CANNED_SOURCE: dict[str, str] = {
    "source_url": "https://example.com/spec",
    "raw_content": "Voltage: 120 VAC\nAirflow: 800 CFM",
}


@pytest.fixture
def client(monkeypatch: pytest.MonkeyPatch) -> TestClient:
    """A TestClient wired up with offline scraper + extractor doubles."""
    monkeypatch.setattr(api, "_scraper", FakeScraper([_CANNED_SOURCE]))
    monkeypatch.setattr(api, "_extractor", FakeExtractor())
    with TestClient(api.app) as test_client:
        yield test_client


def _make_pdf(text: str) -> bytes:
    """Render *text* into an in-memory PDF (same approach as test_pipeline)."""
    document = fitz.open()
    page = document.new_page()
    page.insert_text((72, 72), text)
    return document.tobytes()


def _csv_bytes(rows: list[list[str]]) -> bytes:
    """Serialize *rows* under the canonical batch header."""
    return _csv_bytes_with_headers(
        ["Manufacturer", "Part_Number", "Category"], rows
    )


def _csv_bytes_with_headers(headers: list[str], rows: list[list[str]]) -> bytes:
    """Serialize *rows* under an arbitrary *headers* row."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8")


def _xlsx_bytes(rows: list[list[str]]) -> bytes:
    """Serialize *rows* under the canonical batch header as an .xlsx file."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Manufacturer", "Part_Number", "Category"])
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# POST /api/v1/enrich/single
# ---------------------------------------------------------------------------


def test_single_enrich_json_payload(client: TestClient) -> None:
    response = client.post(
        "/api/v1/enrich/single",
        json={
            "manufacturer_name": "Honeywell",
            "part_number": "TH6320U2008",
            "raw_description": "24 VAC thermostat, 800 CFM airflow.",
            "category": "HVAC",
        },
    )
    assert response.status_code == 200
    body = response.json()

    assert body["sku_id"] == "Honeywell-TH6320U2008"
    assert body["category"] == "HVAC"
    assert len(body["enriched_attributes"]) == 2

    first = body["enriched_attributes"][0]
    assert first["field_name"] == "voltage"
    assert first["normalized_value"] == "120"
    assert first["unit"] == "V"
    # source_url is stamped from the scraped source, not invented by the LLM.
    assert first["source_url"] == "https://example.com/spec"

    assert 0.0 <= body["overall_confidence"] <= 1.0
    assert body["processing_time_ms"] >= 0.0
    assert body["estimated_cost_usd"] >= 0.0


def test_single_enrich_with_pdf_upload(client: TestClient) -> None:
    extractor = api._extractor  # the fake wired in by the fixture
    pdf_bytes = _make_pdf("Voltage: 24 VAC\nAirflow: 800 CFM")

    response = client.post(
        "/api/v1/enrich/single",
        data={
            "manufacturer_name": "Honeywell",
            "part_number": "TH6320U2008",
            "category": "HVAC",
        },
        files={"file": ("datasheet.pdf", pdf_bytes, "application/pdf")},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["sku_id"] == "Honeywell-TH6320U2008"
    assert len(body["enriched_attributes"]) == 2
    # The uploaded PDF text must have reached the extractor, and its source
    # URL must be the upload marker, not a web source.
    assert "24 VAC" in extractor.last_raw_text
    assert "800 CFM" in extractor.last_raw_text
    assert extractor.last_source_url == "upload://datasheet.pdf"
    assert all(
        attribute["source_url"] == "upload://datasheet.pdf"
        for attribute in body["enriched_attributes"]
    )


def test_single_enrich_invalid_category_rejected(client: TestClient) -> None:
    response = client.post(
        "/api/v1/enrich/single",
        json={
            "manufacturer_name": "Honeywell",
            "part_number": "TH6320U2008",
            "category": "Automotive",
        },
    )
    assert response.status_code == 422


def test_single_enrich_unsupported_content_type_rejected(client: TestClient) -> None:
    response = client.post(
        "/api/v1/enrich/single",
        content=b"manufacturer_name=Honeywell",
        headers={"Content-Type": "text/plain"},
    )
    assert response.status_code == 415


def test_single_enrich_missing_api_key_returns_503(monkeypatch: pytest.MonkeyPatch) -> None:
    """With no Groq key configured, extraction degrades to HTTP 503."""
    monkeypatch.setattr(api, "_scraper", FakeScraper([_CANNED_SOURCE]))

    class BrokenExtractor:
        def __init__(self, *args: object, **kwargs: object) -> None:
            raise ValueError("GROQ_API_KEY is required. Set it in .env or pass api_key=")

    monkeypatch.setattr(api, "_extractor", None)
    monkeypatch.setattr(api, "StructuredExtractor", BrokenExtractor)

    with TestClient(api.app) as test_client:
        response = test_client.post(
            "/api/v1/enrich/single",
            json={
                "manufacturer_name": "Honeywell",
                "part_number": "TH6320U2008",
                "category": "HVAC",
            },
        )
    assert response.status_code == 503


# ---------------------------------------------------------------------------
# POST /api/v1/enrich/batch
# ---------------------------------------------------------------------------


def test_batch_enrich_csv(client: TestClient) -> None:
    content = _csv_bytes(
        [["Honeywell", "TH6320U2008", "HVAC"], ["Trane", "XV18", "HVAC"]]
    )
    response = client.post(
        "/api/v1/enrich/batch",
        params={"sync": "true"},
        files={"file": ("batch.csv", content, "text/csv")},
    )
    assert response.status_code == 200
    body = response.json()

    assert body["total"] == 2
    assert body["succeeded"] == 2
    assert body["failed"] == 0
    skus = {result["sku_id"] for result in body["results"]}
    assert skus == {"Honeywell-TH6320U2008", "Trane-XV18"}
    assert all(result["status"] == "success" for result in body["results"])


def test_batch_enrich_async_polling(client: TestClient) -> None:
    """Async 202 Accepted endpoint returns job_id and allows status polling via GET."""
    content = _csv_bytes(
        [["Honeywell", "TH6320U2008", "HVAC"], ["Trane", "XV18", "HVAC"]]
    )
    response = client.post(
        "/api/v1/enrich/batch",
        files={"file": ("batch.csv", content, "text/csv")},
    )
    assert response.status_code == 200 or response.status_code == 202
    body = response.json()
    assert "job_id" in body
    assert body["total_rows"] == 2

    job_id = body["job_id"]
    status_resp = client.get(f"/api/v1/enrich/batch/{job_id}/status")
    assert status_resp.status_code == 200
    status_body = status_resp.json()
    assert status_body["job_id"] == job_id
    assert status_body["total_rows"] == 2


def test_batch_enrich_flexible_manufacturer_headers(client: TestClient) -> None:
    """Lowercase/spaced header variants still yield real manufacturer SKUs
    (no "UNKNOWN-" prefix)."""
    content = _csv_bytes_with_headers(
        ["manufacturer", "Part Number", "category"],
        [["Siemens", "3RT2026-1BB40", "Electrical"]],
    )
    response = client.post(
        "/api/v1/enrich/batch",
        params={"sync": "true"},
        files={"file": ("batch.csv", content, "text/csv")},
    )
    assert response.status_code == 200
    assert response.json()["results"][0]["sku_id"] == "Siemens-3RT2026-1BB40"


def test_batch_enrich_mfr_and_brand_headers(client: TestClient) -> None:
    """``Mfr``/``Brand`` columns map to the manufacturer field."""
    for header in ("Mfr", "Brand"):
        content = _csv_bytes_with_headers(
            [header, "Part_Number", "Category"],
            [["Schneider Electric", "LC1D09", "Electrical"]],
        )
        response = client.post(
            "/api/v1/enrich/batch",
            params={"sync": "true"},
            files={"file": ("batch.csv", content, "text/csv")},
        )
        assert response.status_code == 200
        assert (
            response.json()["results"][0]["sku_id"]
            == "Schneider Electric-LC1D09"
        )


def test_batch_enrich_official_hackathon_headers(client: TestClient) -> None:
    """The official hackathon CSV headers (``Mfg_Part_Num``/``Part_Manuf``/
    ``Part_Desc``, no category column) enrich successfully; the missing
    category defaults to ``"General"``."""
    content = _csv_bytes_with_headers(
        ["Mfg_Part_Num", "Part_Manuf", "Part_Desc"],
        [["TH6320U2008", "Honeywell", "24 VAC thermostat, 800 CFM airflow"]],
    )
    response = client.post(
        "/api/v1/enrich/batch",
        params={"sync": "true"},
        files={"file": ("Unihack_Sample_Dataset_-_Input.csv", content, "text/csv")},
    )
    assert response.status_code == 200
    body = response.json()

    assert body["total"] == 1
    assert body["succeeded"] == 1
    assert body["failed"] == 0
    result = body["results"][0]
    assert result["status"] == "success"
    assert result["sku_id"] == "Honeywell-TH6320U2008"
    assert result["manufacturer_name"] == "Honeywell"
    assert result["part_number"] == "TH6320U2008"
    assert result["category"] == "General"  # no category column -> defaulted
    assert len(result["enriched_attributes"]) == 2


def test_batch_enrich_strips_manufacturer_parenthetical_suffix(client: TestClient) -> None:
    """Trailing code parentheses on manufacturer names are stripped before
    enrichment ("Freud Inc (2435)" -> "Freud Inc")."""
    content = _csv_bytes_with_headers(
        ["Mfg_Part_Num", "Part_Manuf"],
        [["D0700", "Freud Inc (2435)"], ["MIRUS-8", "Mirka Abrasives Inc (MIRUS)"]],
    )
    response = client.post(
        "/api/v1/enrich/batch",
        params={"sync": "true"},
        files={"file": ("batch.csv", content, "text/csv")},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["succeeded"] == 2
    manufacturers = [result["manufacturer_name"] for result in body["results"]]
    assert manufacturers == ["Freud Inc", "Mirka Abrasives Inc"]
    skus = {result["sku_id"] for result in body["results"]}
    assert skus == {"Freud Inc-D0700", "Mirka Abrasives Inc-MIRUS-8"}
    assert all(result["category"] == "General" for result in body["results"])


def test_batch_enrich_excel(client: TestClient) -> None:
    content = _xlsx_bytes([["Honeywell", "TH6320U2008", "HVAC"]])
    response = client.post(
        "/api/v1/enrich/batch",
        params={"sync": "true"},
        files={
            "file": (
                "batch.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 1
    assert body["succeeded"] == 1
    assert body["results"][0]["sku_id"] == "Honeywell-TH6320U2008"


def test_batch_enrich_reports_partial_failure(client: TestClient) -> None:
    content = _csv_bytes(
        [
            ["Honeywell", "TH6320U2008", "HVAC"],
            ["Trane", "XV18", "Automotive"],  # invalid category -> row error
        ]
    )
    response = client.post(
        "/api/v1/enrich/batch",
        params={"sync": "true"},
        files={"file": ("batch.csv", content, "text/csv")},
    )
    assert response.status_code == 200
    body = response.json()

    assert body["total"] == 2
    assert body["succeeded"] == 1
    assert body["failed"] == 1
    error_rows = [result for result in body["results"] if result["status"] == "error"]
    assert len(error_rows) == 1
    assert error_rows[0]["manufacturer_name"] == "Trane"
    assert error_rows[0]["part_number"] == "XV18"
    assert error_rows[0]["error"]


def test_batch_missing_api_key_returns_503(monkeypatch: pytest.MonkeyPatch) -> None:
    """A batch run fails fast with 503 when no Groq key is configured."""
    monkeypatch.setattr(api, "_scraper", FakeScraper([_CANNED_SOURCE]))

    class BrokenExtractor:
        def __init__(self, *args: object, **kwargs: object) -> None:
            raise ValueError("GROQ_API_KEY is required. Set it in .env or pass api_key=")

    monkeypatch.setattr(api, "_extractor", None)
    monkeypatch.setattr(api, "StructuredExtractor", BrokenExtractor)

    content = _csv_bytes([["Honeywell", "TH6320U2008", "HVAC"]])
    with TestClient(api.app) as test_client:
        response = test_client.post(
            "/api/v1/enrich/batch",
            params={"sync": "true"},
            files={"file": ("batch.csv", content, "text/csv")},
        )
    assert response.status_code == 503


def test_batch_rejects_unsupported_extension(client: TestClient) -> None:
    response = client.post(
        "/api/v1/enrich/batch",
        files={"file": ("batch.txt", b"Manufacturer,Part_Number,Category\n", "text/plain")},
    )
    assert response.status_code == 415


def test_batch_rejects_empty_upload(client: TestClient) -> None:
    response = client.post(
        "/api/v1/enrich/batch",
        files={"file": ("batch.csv", b"", "text/csv")},
    )
    assert response.status_code == 400


def test_batch_rejects_over_2000_rows(client: TestClient) -> None:
    """The 2,000-row cap (raised for the 1,000-row hackathon dataset) rejects
    oversized uploads before any enrichment starts."""
    content = _csv_bytes_with_headers(
        ["Mfg_Part_Num", "Part_Manuf"],
        [[f"P{idx}", f"Maker {idx}"] for idx in range(2001)],
    )
    response = client.post(
        "/api/v1/enrich/batch",
        files={"file": ("batch.csv", content, "text/csv")},
    )
    assert response.status_code == 422
    assert "2000" in response.json()["detail"]


def test_chunk_rows_equal_splits() -> None:
    """Rows split into contiguous, roughly-equal shards."""
    assert api._chunk_rows(["a", "b", "c", "d"], 2) == [["a", "b"], ["c", "d"]]
    assert api._chunk_rows(["a", "b", "c"], 1) == [["a", "b", "c"]]
    assert api._chunk_rows(["a", "b", "c"], 5) == [["a"], ["b"], ["c"]]
    assert api._chunk_rows([], 3) == []
    assert api._chunk_rows(["a"], 0) == []


class SpyExtractor(FakeExtractor):
    """FakeExtractor that records the ``api_key`` it was constructed with."""

    instances: list[SpyExtractor] = []

    def __init__(self, api_key: str | None = None, **kwargs: object) -> None:
        super().__init__()
        self.api_key = api_key
        SpyExtractor.instances.append(self)


def test_batch_shards_rows_across_api_keys(monkeypatch: pytest.MonkeyPatch) -> None:
    """With N configured Groq keys, the catalog is split into N contiguous
    shards — each bound to its own key's extractor — and the merged results
    keep the ORIGINAL row order."""
    SpyExtractor.instances = []
    monkeypatch.setattr(api, "_extractor", None)
    monkeypatch.setattr(api, "_extractors_by_key", None)
    monkeypatch.setattr(api, "StructuredExtractor", SpyExtractor)
    monkeypatch.setattr(api, "_scraper", FakeScraper([_CANNED_SOURCE]))
    monkeypatch.setattr(api.settings, "groq_api_keys", "key-a,key-b")

    content = _csv_bytes([[f"Maker {idx}", f"PN-{idx}", "HVAC"] for idx in range(4)])
    with TestClient(api.app) as test_client:
        response = test_client.post(
            "/api/v1/enrich/batch",
            params={"sync": "true"},
            files={"file": ("batch.csv", content, "text/csv")},
        )

    assert response.status_code == 200
    body = response.json()
    assert body["succeeded"] == 4
    # Exactly one extractor per configured key.
    assert [ext.api_key for ext in SpyExtractor.instances] == ["key-a", "key-b"]
    # Shards are merged back in original row order.
    assert [result["sku_id"] for result in body["results"]] == [
        "Maker 0-PN-0",
        "Maker 1-PN-1",
        "Maker 2-PN-2",
        "Maker 3-PN-3",
    ]


# ---------------------------------------------------------------------------
# POST /api/v1/export/excel
# ---------------------------------------------------------------------------


def _sample_products() -> list[dict[str, object]]:
    return [
        {
            "sku_id": "Honeywell-TH6320U2008",
            "category": "HVAC",
            "manufacturer_name": "Honeywell",
            "part_number": "TH6320U2008",
            "overall_confidence": 0.925,
            "processing_time_ms": 12.5,
            "estimated_cost_usd": 0.001,
            "enriched_attributes": [
                {
                    "field_name": "voltage",
                    "raw_value": "120 VAC",
                    "normalized_value": "120",
                    "unit": "V",
                    "confidence_score": 0.95,
                    "source_url": "https://example.com/spec",
                }
            ],
        }
    ]


def test_export_excel_generates_workbook(client: TestClient) -> None:
    response = client.post(
        "/api/v1/export/excel",
        params={"data": json.dumps(_sample_products())},
    )
    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    assert response.headers["content-disposition"].startswith("attachment")

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    assert "Enriched Catalog" in workbook.sheetnames

    catalog_sheet = workbook["Enriched Catalog"]
    headers = [cell.value for cell in next(catalog_sheet.iter_rows(min_row=1, max_row=1))]
    assert len(headers) == 252
    assert headers[0] == "SKU - MY_PART_NUMBER"
    assert headers[1] == "MANUFACTURER_NAME"
    assert headers[4] == "PART_NUMBER"

    product_rows = list(catalog_sheet.iter_rows(min_row=2, values_only=True))
    assert len(product_rows) == 1
    assert product_rows[0][0] == "Honeywell-TH6320U2008"
    assert product_rows[0][1] == "Honeywell"
    assert product_rows[0][4] == "TH6320U2008"


def test_export_excel_accepts_json_body(client: TestClient) -> None:
    response = client.post(
        "/api/v1/export/excel",
        content=json.dumps(_sample_products()),
        headers={"Content-Type": "application/json"},
    )
    assert response.status_code == 200
    assert len(response.content) > 0


def test_export_excel_rejects_invalid_json(client: TestClient) -> None:
    response = client.post("/api/v1/export/excel", params={"data": "{not json"})
    assert response.status_code == 400


def test_export_excel_rejects_non_list_data(client: TestClient) -> None:
    response = client.post(
        "/api/v1/export/excel",
        params={"data": json.dumps({"sku_id": "not-a-list"})},
    )
    assert response.status_code == 422


def test_export_excel_rejects_missing_data(client: TestClient) -> None:
    response = client.post("/api/v1/export/excel")
    assert response.status_code == 422


def test_export_excel_rejects_get_method(client: TestClient) -> None:
    """Export is a POST endpoint — GET must be rejected (405)."""
    response = client.get(
        "/api/v1/export/excel",
        params={"data": json.dumps(_sample_products())},
    )
    assert response.status_code == 405


def test_batch_enrich_hits_cache_on_rerun(client: TestClient) -> None:
    """Re-running the same batch dataset immediately hits SQLite enrichment cache."""
    csv_bytes = b"Manufacturer,Part Number,Description\nSquare D,QO120,Circuit Breaker 20A\n"

    # First run: extracts via extractor and populates cache
    r1 = client.post(
        "/api/v1/enrich/batch",
        params={"sync": "true"},
        files={"file": ("batch.csv", csv_bytes, "text/csv")},
    )
    assert r1.status_code == 200
    res1 = r1.json()
    assert res1["succeeded"] == 1
    assert len(res1["results"][0]["enriched_attributes"]) == 2

    # Second run: must hit pre-execution cache guard with identical attributes
    r2 = client.post(
        "/api/v1/enrich/batch",
        params={"sync": "true"},
        files={"file": ("batch.csv", csv_bytes, "text/csv")},
    )
    assert r2.status_code == 200
    res2 = r2.json()
    assert res2["succeeded"] == 1
    assert len(res2["results"][0]["enriched_attributes"]) == 2
    # Verify cached attributes match
    assert res2["results"][0]["enriched_attributes"][0]["field_name"] == "voltage"

