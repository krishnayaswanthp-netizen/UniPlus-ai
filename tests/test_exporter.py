"""
tests/test_exporter.py
Stage 12 Verification: Tests for JSON export, 4-tab openpyxl Excel export, and manual review queue population.
"""

import json
import os
import pytest
from openpyxl import load_workbook

from app.schemas.product import (
    ProductRecord,
    ProductIdentity,
    RawInputData,
    RowStatus,
    AttributeValue,
    ExtractionSource,
)
from app.services.exporter import CatalogExporter


def _make_sample_record(row_id=1, status=RowStatus.COMPLETED, validity=True):
    identity = ProductIdentity(
        row_id=row_id,
        mfg_part_number="775L",
        manufacturer="3M",
        raw_description="3M Disc",
        category="Abrasives",
    )
    record = ProductRecord(
        identity=identity, raw_data=RawInputData(original_row_index=0)
    )
    record.status = status
    record.quality.validity = validity
    record.quality.completeness = 1.0
    record.quality.overall_confidence = 0.95
    record.attributes["voltage"] = AttributeValue(
        field_name="voltage",
        raw_value="24V",
        normalized_value="24",
        unit="V",
        confidence=0.99,
        source=ExtractionSource.REGEX,
    )
    return record


def test_export_to_json(tmp_path):
    json_file = str(tmp_path / "catalog.json")
    rec = _make_sample_record()

    out_path = CatalogExporter.export_to_json([rec], json_file)
    assert os.path.exists(out_path)

    with open(out_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    assert len(data) == 1
    assert data[0]["identity"]["sku_id"] == "3M-775L"


def test_export_to_excel_creates_four_tabs(tmp_path):
    excel_file = str(tmp_path / "catalog.xlsx")
    rec1 = _make_sample_record(row_id=1, status=RowStatus.COMPLETED)
    rec2 = _make_sample_record(row_id=2, status=RowStatus.MANUAL_REVIEW, validity=False)
    rec2.quality.validation_flags = ["INVALID_VOLTAGE_NEGATIVE"]

    snapshot = {
        "completed_rows": 1,
        "llm_bypass_ratio": 1.0,
        "total_tokens_consumed": 0,
        "rows_per_second": 10.5,
    }

    out_path = CatalogExporter.export_to_excel([rec1, rec2], snapshot, excel_file)
    assert os.path.exists(out_path)

    wb = load_workbook(out_path)
    sheet_names = wb.sheetnames

    assert "Enriched Catalog" in sheet_names
    assert "Lineage Audit" in sheet_names
    assert "Judge Telemetry" in sheet_names
    assert "Manual Review Queue" in sheet_names

    # Check Manual Review Queue has row 2
    ws_review = wb["Manual Review Queue"]
    assert ws_review.max_row >= 2
    assert ws_review.cell(row=2, column=1).value == 2

    # Check Enriched Catalog has 252 columns
    ws_catalog = wb["Enriched Catalog"]
    headers = [cell.value for cell in next(ws_catalog.iter_rows(min_row=1, max_row=1))]
    assert len(headers) == 252


def test_export_excel_bytes_unilog_schema():
    rec = _make_sample_record()
    b = CatalogExporter.export_excel_bytes([rec])
    assert isinstance(b, bytes)
    assert len(b) > 0

