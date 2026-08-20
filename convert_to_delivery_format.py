#!/usr/bin/env python3
"""
convert_to_delivery_format.py
Standalone conversion utility mapping database exports / now.xlsx into the exact
252-column flat Unilog Delivery Schema (Unihack_Submission_Output.csv & .xlsx).

Usage:
    python convert_to_delivery_format.py [input_file.xlsx]
"""

from __future__ import annotations

import csv
import json
import os
import sqlite3
import sys
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook

# Import the official 252-column schema and mapping functions from app.services.exporter
try:
    from app.services.exporter import (
        UNILOG_COLUMNS,
        CatalogExporter,
        clean_attribute_label,
        map_record_to_unilog_row,
    )
except ImportError:
    # Ensure current directory is on python path
    sys.path.insert(0, os.path.abspath(os.path.dirname(__file__)))
    from app.services.exporter import (
        UNILOG_COLUMNS,
        CatalogExporter,
        clean_attribute_label,
        map_record_to_unilog_row,
    )


def load_records_from_excel(file_path: str) -> list[dict[str, Any]]:
    """Parse existing now.xlsx (relational Products/Attributes or flat catalog)."""
    wb = load_workbook(file_path, data_only=True)
    sheet_names = wb.sheetnames

    # Case A: Relational sheets 'Products' and 'Attributes'
    if "Products" in sheet_names and "Attributes" in sheet_names:
        products_ws = wb["Products"]
        attributes_ws = wb["Attributes"]

        # Read Products headers & rows
        p_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in products_ws[1]]
        products_map: dict[str, dict[str, Any]] = {}
        for row in products_ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            p_dict = dict(zip(p_headers, row))
            sku = str(p_dict.get("sku_id") or "").strip()
            if sku:
                products_map[sku] = {
                    "sku_id": sku,
                    "category": p_dict.get("category") or "General",
                    "manufacturer_name": p_dict.get("manufacturer_name") or "",
                    "part_number": p_dict.get("part_number") or "",
                    "overall_confidence": p_dict.get("overall_confidence") or 0.95,
                    "enriched_attributes": [],
                }

        # Read Attributes headers & rows
        a_headers = [str(cell.value).strip() if cell.value is not None else "" for cell in attributes_ws[1]]
        for row in attributes_ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            a_dict = dict(zip(a_headers, row))
            sku = str(a_dict.get("sku_id") or "").strip()
            if sku in products_map:
                products_map[sku]["enriched_attributes"].append({
                    "field_name": a_dict.get("field_name") or "",
                    "raw_value": a_dict.get("raw_value") or "",
                    "normalized_value": a_dict.get("normalized_value") or a_dict.get("raw_value") or "",
                    "unit": a_dict.get("unit") or None,
                    "confidence_score": a_dict.get("confidence_score") or 0.95,
                    "source_url": a_dict.get("source_url") or "",
                })

        return list(products_map.values())

    # Case B: Single sheet containing headers
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        records.append(dict(zip(headers, row)))
    return records


def load_records_from_sqlite(db_path: str = "unipulse_checkpoint.db") -> list[dict[str, Any]]:
    """Retrieve checkpointed records from SQLite database if now.xlsx is not found."""
    if not os.path.exists(db_path):
        return []

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT record_json FROM checkpoints WHERE status = 'COMPLETED'")
    rows = cursor.fetchall()
    conn.close()

    records: list[dict[str, Any]] = []
    for (json_str,) in rows:
        try:
            d = json.loads(json_str)
            identity = d.get("identity", {})
            mfg = identity.get("manufacturer") or ""
            part = identity.get("mfg_part_number") or ""
            cat = identity.get("category") or "General"
            sku = identity.get("sku_id") or f"{mfg}-{part}"
            attrs = []
            for k, v in d.get("attributes", {}).items():
                attrs.append({
                    "field_name": k,
                    "raw_value": v.get("raw_value", ""),
                    "normalized_value": v.get("normalized_value", v.get("raw_value", "")),
                    "unit": v.get("unit"),
                    "confidence_score": v.get("confidence", 0.95),
                    "source_url": v.get("evidence_snippet", ""),
                })
            records.append({
                "sku_id": sku,
                "category": cat,
                "manufacturer_name": mfg,
                "part_number": part,
                "enriched_attributes": attrs,
            })
        except Exception:
            continue
    return records


def create_sample_now_xlsx(output_path: str = "now.xlsx") -> None:
    """Create a sample relational now.xlsx file if no input file exists."""
    wb = Workbook()
    ws_prod = wb.active
    ws_prod.title = "Products"
    ws_prod.append([
        "sku_id", "category", "manufacturer_name", "part_number",
        "overall_confidence", "processing_time_ms", "estimated_cost_usd", "attribute_count"
    ])
    ws_prod.append([
        "Honeywell-TH6320U2008", "HVAC", "Honeywell", "TH6320U2008",
        0.98, 120.5, 0.0004, 3
    ])
    ws_prod.append([
        "3M-775L", "Abrasives", "3M", "775L",
        0.96, 95.2, 0.0003, 2
    ])

    ws_attr = wb.create_sheet("Attributes")
    ws_attr.append([
        "sku_id", "field_name", "raw_value", "normalized_value", "unit", "confidence_score", "source_url"
    ])
    ws_attr.append(["Honeywell-TH6320U2008", "voltage", "24VAC", "24", "VAC", 0.99, "https://customer.resideo.com/resources"])
    ws_attr.append(["Honeywell-TH6320U2008", "stages", "3 Heat / 2 Cool", "3 Heat / 2 Cool", None, 0.95, "https://customer.resideo.com/resources"])
    ws_attr.append(["Honeywell-TH6320U2008", "display_size", "6.89 sq in", "6.89", "sq in", 0.97, "https://customer.resideo.com/resources"])
    ws_attr.append(["3M-775L", "grit", "80+", "80+", None, 0.99, "https://www.3m.com/abrasives"])
    ws_attr.append(["3M-775L", "diameter", "5 in", "5", "in", 0.98, "https://www.3m.com/abrasives"])

    wb.save(output_path)
    print(f"[*] Generated default sample source: {output_path}")


def convert(input_xlsx: str = "now.xlsx") -> tuple[str, str, int]:
    """Execute the conversion from now.xlsx / SQLite to Unihack delivery format."""
    print(f"[*] Locating product dataset (Target: {input_xlsx})...")

    records: list[dict[str, Any]] = []
    if os.path.exists(input_xlsx):
        records = load_records_from_excel(input_xlsx)
        print(f"[*] Loaded {len(records)} records from {input_xlsx}")
    else:
        # Check SQLite DB
        records = load_records_from_sqlite("unipulse_checkpoint.db")
        if records:
            print(f"[*] Loaded {len(records)} records from unipulse_checkpoint.db")
        else:
            print(f"[*] No existing {input_xlsx} or populated database found. Generating sample data...")
            create_sample_now_xlsx(input_xlsx)
            records = load_records_from_excel(input_xlsx)

    if not records:
        print("[!] No records found to convert.")
        return "", "", 0

    # Map all records to 252-column flat Unilog rows
    mapped_rows: list[dict[str, Any]] = [map_record_to_unilog_row(r) for r in records]

    # 1. Write CSV
    csv_output = "Unihack_Submission_Output.csv"
    with open(csv_output, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=UNILOG_COLUMNS)
        writer.writeheader()
        writer.writerows(mapped_rows)
    print(f"[OK] Saved CSV Output: {csv_output} (Rows: {len(mapped_rows)}, Columns: {len(UNILOG_COLUMNS)})")

    # 2. Write XLSX
    xlsx_output = "Unihack_Submission_Output.xlsx"
    wb = CatalogExporter.build_unilog_workbook(mapped_rows)
    wb.save(xlsx_output)
    print(f"[OK] Saved XLSX Output: {xlsx_output} (Rows: {len(mapped_rows)}, Columns: {len(UNILOG_COLUMNS)})")

    return csv_output, xlsx_output, len(mapped_rows)


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else "now.xlsx"
    csv_file, xlsx_file, count = convert(target)
    print("\n=======================================================")
    print(" UNILOG 252-COLUMN DELIVERY CONVERSION COMPLETED")
    print("=======================================================")
    print(f" Total Products Processed : {count}")
    print(f" Schema Column Width      : {len(UNILOG_COLUMNS)} columns")
    print(f" CSV Delivery File        : {csv_file}")
    print(f" XLSX Delivery File       : {xlsx_file}")
    print("=======================================================\n")
