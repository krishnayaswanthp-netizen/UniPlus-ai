"""
app/services/exporter.py
Stage 12: Enterprise 252-Column Unilog Delivery Schema & Multi-Tab Excel/JSON Exporter.

Produces:
1. Exact flat 252-column Unilog delivery schema (matching Unihack Expected Output format).
2. Multi-tab workbooks with Enriched Catalog, Lineage Audit, Judge Telemetry, and Manual Review.
3. High-performance in-memory XLSX byte streams for FastAPI endpoints.
"""

from __future__ import annotations

import io
import json
import re
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.enrichment import IndustrialAttribute, ProductEnrichmentResponse
from app.schemas.product import ProductRecord, RowStatus
from app.services.provenance import ProvenanceMerger

# ---------------------------------------------------------------------------
# 252-Column Unilog Delivery Schema Specification
# ---------------------------------------------------------------------------

_CORE_UNILOG_COLUMNS: list[str] = [
    # 1-10: Identifiers & Names
    "SKU - MY_PART_NUMBER",
    "MANUFACTURER_NAME",
    "MANUFACTURER_PART_NUMBER",
    "Mfg_Part_Num",
    "PART_NUMBER",
    "Product Name",
    "Classpath",
    "BRAND",
    "SERIES",
    "MODEL_NUMBER",
    # 11-20: Descriptions & Content
    "SHORT_DESC",
    "LONG_DESC1",
    "LONG_DESC2",
    "LONG_DESC3",
    "LONG_DESC4",
    "MOBILE_DESC",
    "FEATURES_BENEFITS",
    "APPLICATIONS",
    "INCLUDES",
    "STANDARDS_COMPLIANCE",
    # 21-30: URLs & Media
    "MFR URL",
    "Ref URL 1",
    "Ref URL 2",
    "Ref URL 3",
    "Ref URL 4",
    "Ref URL 5",
    "IMAGE URL",
    "IMAGE URL 2",
    "DOCUMENT URL",
    "SPEC_SHEET_URL",
    # 31-40: Categorization & Industry Codes
    "CATEGORY",
    "SUB_CATEGORY",
    "UNSPSC",
    "UPC",
    "GTIN",
    "EAN",
    "STATUS",
    "COUNTRY_OF_ORIGIN",
    "TARIFF_CODE",
    "HAZMAT_CODE",
    # 41-50: Commercial & Packaging
    "PRICE",
    "CURRENCY",
    "PACK_QTY",
    "PACK_UOM",
    "MIN_ORDER_QTY",
    "LEAD_TIME_DAYS",
    "WARRANTY",
    "HAZARDOUS_MATERIAL",
    "PROP_65_WARNING",
    "ROHS_COMPLIANT",
    # 51-60: Certifications & Standards
    "ENERGY_STAR",
    "UL_LISTED",
    "CSA_CERTIFIED",
    "CE_CERTIFIED",
    "ETL_LISTED",
    "NEMA_RATING",
    "IP_RATING",
    "REACH_COMPLIANT",
    "ISO_CERTIFIED",
    "CERTIFICATIONS",
    # 61-70: Physical Dimensions
    "LENGTH",
    "LENGTH_UOM",
    "WIDTH",
    "WIDTH_UOM",
    "HEIGHT",
    "HEIGHT_UOM",
    "OVERALL_DEPTH",
    "DEPTH_UOM",
    "WEIGHT",
    "WEIGHT_UOM",
    # 71-80: Materials & Appearance
    "COLOR",
    "MATERIAL",
    "FINISH",
    "MOUNTING_TYPE",
    "CONNECTION_TYPE",
    "PORT_SIZE",
    "THREAD_SIZE",
    "ENCLOSURE_TYPE",
    "HOUSING_MATERIAL",
    "BODY_STYLE",
    # 81-90: Electrical Specifications
    "VOLTAGE",
    "VOLTAGE_UOM",
    "VOLTAGE_TYPE",
    "AMPERAGE",
    "AMPERAGE_UOM",
    "WATTAGE",
    "WATTAGE_UOM",
    "FREQUENCY",
    "FREQUENCY_UOM",
    "PHASE",
    # 91-102: Mechanical & Environmental Properties
    "HORSEPOWER",
    "PRESSURE_RATING",
    "PRESSURE_UOM",
    "FLOW_RATE",
    "FLOW_RATE_UOM",
    "TEMPERATURE_MIN",
    "TEMPERATURE_MAX",
    "TEMPERATURE_UOM",
    "EFFICIENCY_RATING",
    "NOISE_LEVEL_DBA",
    "OPERATING_SPEED_RPM",
    "CUSTOM_NOTES",
]

# Generate 50 dynamic attribute triplets: ATTRIBUTE_LABEL 1..50, ATTRIBUTE_VALUE 1..50, ATTRIBUTE_UOM 1..50
_DYNAMIC_TRIPLETS: list[str] = []
for i in range(1, 51):
    _DYNAMIC_TRIPLETS.extend([
        f"ATTRIBUTE_LABEL {i}",
        f"ATTRIBUTE_VALUE {i}",
        f"ATTRIBUTE_UOM {i}",
    ])

UNILOG_COLUMNS: list[str] = _CORE_UNILOG_COLUMNS + _DYNAMIC_TRIPLETS
assert len(UNILOG_COLUMNS) == 252, f"Expected exactly 252 columns, got {len(UNILOG_COLUMNS)}"


def clean_attribute_label(name: str) -> str:
    """Format raw attribute field names into clean title-case labels."""
    if not name:
        return ""
    s = str(name).strip()
    # Strip common unit suffix tags e.g. _in, _mm, _psi, _vac, _v, _hz, _cfm, _degc, _f
    s = re.sub(
        r"_(?:in|mm|cm|m|ft|psi|bar|kpa|vac|vdc|v|a|amp|amps|w|kw|hz|khz|mhz|cfm|gpm|degc|degf|c|f|lb|lbs|kg|g|oz)$",
        "",
        s,
        flags=re.IGNORECASE,
    )
    s = s.replace("_", " ").strip()
    return s.title()


def map_record_to_unilog_row(item: Any) -> dict[str, Any]:
    """Transform a ProductRecord, ProductEnrichmentResponse, or dict into the 252 Unilog columns."""
    row: dict[str, Any] = {col: "" for col in UNILOG_COLUMNS}

    # Extract common identity attributes
    mfg = ""
    part = ""
    category = "General"
    sku_id = ""
    attributes: list[dict[str, Any]] = []

    if isinstance(item, ProductRecord):
        mfg = item.identity.manufacturer or ""
        part = item.identity.mfg_part_number or ""
        category = item.identity.category or "General"
        sku_id = item.identity.sku_id or f"{mfg}-{part}"
        for attr_key, attr_val in item.attributes.items():
            attributes.append({
                "field_name": attr_key,
                "raw_value": attr_val.raw_value or "",
                "normalized_value": attr_val.normalized_value or attr_val.raw_value or "",
                "unit": attr_val.unit,
                "confidence_score": attr_val.confidence,
                "source_url": attr_val.evidence_snippet or "",
            })
    elif isinstance(item, ProductEnrichmentResponse):
        sku_id = item.sku_id or ""
        category = item.category or "General"
        if "-" in sku_id:
            parts = sku_id.split("-", 1)
            mfg = parts[0]
            part = parts[1]
        for attr in item.enriched_attributes:
            attributes.append({
                "field_name": attr.field_name,
                "raw_value": attr.raw_value,
                "normalized_value": attr.normalized_value or attr.raw_value,
                "unit": attr.unit,
                "confidence_score": attr.confidence_score,
                "source_url": attr.source_url or "",
            })
    elif isinstance(item, dict):
        mfg = item.get("manufacturer_name") or item.get("manufacturer") or item.get("mfg") or ""
        part = item.get("part_number") or item.get("part_num") or item.get("mfg_part_number") or ""
        category = item.get("category") or "General"
        sku_id = item.get("sku_id") or (f"{mfg}-{part}" if mfg and part else "")
        if not mfg and sku_id and "-" in sku_id:
            parts = sku_id.split("-", 1)
            mfg = parts[0]
            part = parts[1]
        raw_attrs = item.get("enriched_attributes") or item.get("attributes") or []
        if isinstance(raw_attrs, list):
            for a in raw_attrs:
                if isinstance(a, dict):
                    attributes.append(a)
                elif hasattr(a, "field_name"):
                    attributes.append({
                        "field_name": getattr(a, "field_name", ""),
                        "raw_value": getattr(a, "raw_value", ""),
                        "normalized_value": getattr(a, "normalized_value", getattr(a, "raw_value", "")),
                        "unit": getattr(a, "unit", None),
                        "confidence_score": getattr(a, "confidence_score", getattr(a, "confidence", 0.95)),
                        "source_url": getattr(a, "source_url", ""),
                    })
        elif isinstance(raw_attrs, dict):
            for k, v in raw_attrs.items():
                if isinstance(v, dict):
                    attributes.append({"field_name": k, **v})
                else:
                    attributes.append({"field_name": k, "raw_value": str(v), "normalized_value": str(v)})

    # Core Identifiers
    row["SKU - MY_PART_NUMBER"] = sku_id or f"{mfg}-{part}".strip("-")
    row["MANUFACTURER_NAME"] = mfg
    row["MANUFACTURER_PART_NUMBER"] = part
    row["Mfg_Part_Num"] = part
    row["PART_NUMBER"] = part
    row["Product Name"] = f"{mfg} {part}".strip()
    row["Classpath"] = f"Industrial & Commercial > {category}"
    row["BRAND"] = mfg
    row["CATEGORY"] = category
    row["STATUS"] = "Active"

    # Descriptions
    row["SHORT_DESC"] = f"{mfg} {part} {category} Component".strip()
    row["MOBILE_DESC"] = f"{mfg}, {category}, {part}".strip(", ")

    # Build LONG_DESC1 and extract URLs
    desc_snippets: list[str] = []
    urls: list[str] = []

    for attr in attributes:
        fname = clean_attribute_label(attr.get("field_name", ""))
        val = str(attr.get("normalized_value") or attr.get("raw_value") or "").strip()
        u = str(attr.get("unit") or "").strip()
        if fname and val:
            desc_snippets.append(f"{fname}: {val} {u}".strip())
        src = str(attr.get("source_url") or "").strip()
        if src.startswith("http://") or src.startswith("https://"):
            if src not in urls:
                urls.append(src)

    if desc_snippets:
        row["LONG_DESC1"] = "; ".join(desc_snippets)

    # Populate URLs
    if urls:
        row["MFR URL"] = urls[0]
        for idx, u in enumerate(urls[:5], start=1):
            row[f"Ref URL {idx}"] = u

    # Physical Dimension Shortcuts & Electrical Mapping
    for attr in attributes:
        fn_raw = str(attr.get("field_name", "")).strip().lower()
        val = str(attr.get("normalized_value") or attr.get("raw_value") or "").strip()
        u = str(attr.get("unit") or "").strip()

        if fn_raw in ("length", "overall_length", "len") and not row["LENGTH"]:
            row["LENGTH"] = val
            row["LENGTH_UOM"] = u or "in"
        elif fn_raw in ("width", "overall_width", "wid") and not row["WIDTH"]:
            row["WIDTH"] = val
            row["WIDTH_UOM"] = u or "in"
        elif fn_raw in ("height", "overall_height", "hgt") and not row["HEIGHT"]:
            row["HEIGHT"] = val
            row["HEIGHT_UOM"] = u or "in"
        elif fn_raw in ("weight", "item_weight", "wt") and not row["WEIGHT"]:
            row["WEIGHT"] = val
            row["WEIGHT_UOM"] = u or "lbs"
        elif fn_raw in ("voltage", "voltage_vac", "voltage_vdc", "operating_voltage") and not row["VOLTAGE"]:
            row["VOLTAGE"] = val
            row["VOLTAGE_UOM"] = u or "V"
        elif fn_raw in ("amperage", "current", "current_rating", "amps") and not row["AMPERAGE"]:
            row["AMPERAGE"] = val
            row["AMPERAGE_UOM"] = u or "A"
        elif fn_raw in ("wattage", "power", "power_rating", "watts") and not row["WATTAGE"]:
            row["WATTAGE"] = val
            row["WATTAGE_UOM"] = u or "W"
        elif fn_raw in ("frequency", "freq") and not row["FREQUENCY"]:
            row["FREQUENCY"] = val
            row["FREQUENCY_UOM"] = u or "Hz"
        elif fn_raw in ("pressure_rating", "pressure", "operating_pressure") and not row["PRESSURE_RATING"]:
            row["PRESSURE_RATING"] = val
            row["PRESSURE_UOM"] = u or "PSI"
        elif fn_raw in ("flow_rate", "airflow", "airflow_cfm") and not row["FLOW_RATE"]:
            row["FLOW_RATE"] = val
            row["FLOW_RATE_UOM"] = u or "CFM"
        elif fn_raw in ("material", "body_material") and not row["MATERIAL"]:
            row["MATERIAL"] = val
        elif fn_raw in ("color",) and not row["COLOR"]:
            row["COLOR"] = val
        elif fn_raw in ("finish",) and not row["FINISH"]:
            row["FINISH"] = val
        elif fn_raw in ("mounting_type", "mounting") and not row["MOUNTING_TYPE"]:
            row["MOUNTING_TYPE"] = val

    # Dynamic Attribute Triplets (1..50)
    for idx, attr in enumerate(attributes[:50], start=1):
        lbl = clean_attribute_label(attr.get("field_name", ""))
        val = str(attr.get("normalized_value") or attr.get("raw_value") or "").strip()
        u = str(attr.get("unit") or "").strip()
        row[f"ATTRIBUTE_LABEL {idx}"] = lbl
        row[f"ATTRIBUTE_VALUE {idx}"] = val
        row[f"ATTRIBUTE_UOM {idx}"] = u

    return row


class CatalogExporter:
    """Enterprise 252-column Unilog delivery schema & multi-tab Excel/JSON export engine."""

    UNILOG_COLUMNS = UNILOG_COLUMNS

    @classmethod
    def build_unilog_workbook(cls, records: List[Any]) -> Workbook:
        """Create an openpyxl Workbook with the flat 252-column 'Enriched Catalog' sheet."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Enriched Catalog"

        # Headers
        ws.append(UNILOG_COLUMNS)

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for item in records:
            mapped = map_record_to_unilog_row(item)
            ws.append([mapped.get(col, "") for col in UNILOG_COLUMNS])

        # Auto-fit width & freeze panes
        for col_idx, col_name in enumerate(UNILOG_COLUMNS, start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max(len(col_name) + 3, 14)
        ws.freeze_panes = "A2"

        return wb

    @classmethod
    def export_excel_bytes(cls, records: List[Any]) -> bytes:
        """Render records directly to in-memory Excel XLSX byte stream."""
        wb = cls.build_unilog_workbook(records)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @classmethod
    def export_to_json(cls, records: List[ProductRecord], file_path: str) -> str:
        """Exports a list of ProductRecords to a formatted JSON file."""
        data = []
        for rec in records:
            dumped = json.loads(rec.model_dump_json())
            dumped["identity"]["sku_id"] = rec.identity.sku_id
            data.append(dumped)
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        return file_path

    @classmethod
    def export_to_excel(
        cls,
        records: List[ProductRecord],
        metrics_snapshot: Dict[str, Any],
        output_path: str,
    ) -> str:
        """Exports ProductRecords and telemetry snapshot into a 4-tab openpyxl Excel workbook:
        1. Enriched Catalog (252-Column Unilog Schema)
        2. Lineage Audit
        3. Judge Telemetry
        4. Manual Review Queue
        """
        wb = cls.build_unilog_workbook(records)

        # Styling
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        review_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

        # -------------------------------------------------------------------
        # TAB 2: Lineage Audit
        # -------------------------------------------------------------------
        ws_audit = wb.create_sheet(title="Lineage Audit")
        export_dicts = [
            ProvenanceMerger.build_provenance_export_dict(rec) for rec in records
        ]

        audit_headers: List[str] = []
        if export_dicts:
            audit_headers = list(export_dicts[0].keys())
            for d in export_dicts[1:]:
                for k in d.keys():
                    if k not in audit_headers:
                        audit_headers.append(k)

        ws_audit.append(audit_headers)
        for cell in ws_audit[1]:
            cell.font = header_font
            cell.fill = header_fill

        for d in export_dicts:
            ws_audit.append([d.get(k, "") for k in audit_headers])

        # -------------------------------------------------------------------
        # TAB 3: Judge Telemetry
        # -------------------------------------------------------------------
        ws_telemetry = wb.create_sheet(title="Judge Telemetry")
        ws_telemetry.append(["Metric / Telemetry Indicator", "Observed Value"])
        for cell in ws_telemetry[1]:
            cell.font = header_font
            cell.fill = header_fill

        for k, v in (metrics_snapshot or {}).items():
            ws_telemetry.append([k.replace("_", " ").title(), str(v)])

        # -------------------------------------------------------------------
        # TAB 4: Manual Review Queue
        # -------------------------------------------------------------------
        ws_review = wb.create_sheet(title="Manual Review Queue")
        review_headers = [
            "Row ID",
            "SKU ID",
            "Manufacturer",
            "Part Number",
            "Category",
            "Status",
            "Validation Flags",
            "Description",
        ]
        ws_review.append(review_headers)
        for cell in ws_review[1]:
            cell.font = header_font
            cell.fill = review_fill

        for rec in records:
            if rec.status == RowStatus.MANUAL_REVIEW or not rec.quality.validity:
                ws_review.append([
                    rec.identity.row_id,
                    rec.identity.sku_id,
                    rec.identity.manufacturer,
                    rec.identity.mfg_part_number,
                    rec.identity.category,
                    rec.status.value,
                    ", ".join(rec.quality.validation_flags),
                    rec.identity.raw_description,
                ])

        wb.save(output_path)
        return output_path
