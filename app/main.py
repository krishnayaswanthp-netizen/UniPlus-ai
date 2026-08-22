"""UniPulse AI — FastAPI application entry point.

Registers the product-enrichment API:

- ``POST /api/v1/enrich/single``  — enrich one product from a JSON payload
  (``ProductEnrichmentRequest``) and/or an uploaded PDF datasheet.
- ``POST /api/v1/enrich/batch``   — enrich many products from an uploaded
  CSV / Excel file, processed concurrently with a bounded semaphore.
- ``POST /api/v1/export/excel``   — render enriched results as a
  downloadable ``.xlsx`` workbook (``openpyxl``).

  The export endpoint accepts the enriched-product array as a JSON request
  body (or, for small payloads, a ``data`` query parameter). It is ``POST``
  because browser clients must not rely on ``GET`` request bodies, and large
  batches exceed URL length limits.

The enrichment pipeline itself lives in :func:`_enrich_single_product` and
reuses the existing services: ``WhitelistedSearchScraper`` (web search +
whitelisted scraping), ``PDFParser`` (uploaded datasheets),
``StructuredExtractor`` (LLM extraction; ``UnitNormalizer`` is applied
inside its post-processing step) and ``UnitNormalizer``.
"""

from __future__ import annotations

import asyncio
import csv
import io
import json
import logging
import math
import re
import time
import uuid
from typing import Any

logger = logging.getLogger(__name__)

from fastapi import FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pydantic import ValidationError

from app.core.config import settings
from app.schemas.enrichment import (
    BatchEnrichmentResponse,
    BatchItemResult,
    IndustrialAttribute,
    ProductEnrichmentRequest,
    ProductEnrichmentResponse,
)
from app.db.checkpoint_store import CheckpointStore
from app.schemas.product import (
    AttributeValue,
    ExtractionSource,
    ProductIdentity,
    ProductRecord,
    RawInputData,
    RowStatus,
)
from app.services.extractor import StructuredExtractor, _MAX_INPUT_CHARS
from app.services.exporter import CatalogExporter
from app.services.parser import PDFParser
from app.services.rate_limiter import AdaptiveRateLimiter
from app.services.scraper import WhitelistedSearchScraper

_checkpoint_store = CheckpointStore(settings.database_path)

app = FastAPI(
    title=settings.app_name,
    version=settings.app_version,
    description="High-throughput B2B product intelligence engine.",
)

# CORS: allow browser-based clients (e.g. the Streamlit dashboard) to call the API.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# Pipeline wiring
# ---------------------------------------------------------------------------

_pdf_parser = PDFParser()
_scraper = WhitelistedSearchScraper()
#: Built lazily because ``StructuredExtractor`` requires a Groq API key at
#: construction time; importing ``app.main`` must not fail without one.
_extractor: StructuredExtractor | None = None

#: Per-key extractors used to shard batch uploads across ``GROQ_API_KEYS``
#: (one extractor per key, built lazily). Distinct from ``_extractor`` so a
#: /single request never collapses batch sharding: batch always builds its
#: own keyed extractors when multiple keys are configured.
_extractors_by_key: list[StructuredExtractor] | None = None

#: URL stamped onto attributes extracted from an uploaded PDF datasheet.
_UPLOAD_SOURCE_PREFIX = "upload://"
#: Fallback label used when a product has no scrape/PDF/description content.
_FALLBACK_SOURCE_URL = "local://user-provided"

#: Upper bound on rows processed per batch upload (guard against runaway files).
#: Raised to 2,000 so official hackathon datasets (1,000 rows) fit comfortably.
_MAX_BATCH_ROWS = 2000

#: Rows processed concurrently inside a single key-shard. Bounded to 2 to keep
#: memory strictly < 250MB and prevent OOM on Render free tier (512MB RAM cap).
_SHARD_CONCURRENCY = 2
#: Stagger between row dispatches inside a shard. Produces steady, non-bursty
#: per-key throughput instead of a wall of simultaneous calls.
_SHARD_DISPATCH_DELAY = 0.1

#: Global cap on concurrent LLM extraction calls, shared across ALL requests
#: (single + batch, every user). Per-shard semaphores bound row-level work
#: inside one upload; this bounds total Groq API concurrency server-wide so
#: two concurrent batches can't double the load on the rate limit.
_LLM_CONCURRENCY = asyncio.Semaphore(8)

#: Global token-budget gate for the /single and /batch LLM calls. Groq's
#: free-tier per-key ceiling is 12,000 TPM; the limiter caps the sliding
#: window at 9,500 (~80%) so concurrent rows stay safely under it even when
#: token estimates run hot. RPM is set high because request concurrency is
#: already bounded by ``_LLM_CONCURRENCY`` — TPM is the binding constraint.
_RATE_LIMITER = AdaptiveRateLimiter(max_rpm=120)

_HEADER_FONT = Font(bold=True)


def _get_extractor() -> StructuredExtractor:
    """Return the shared extractor, building it on first use.

    Raises
    ------
    HTTPException
        With status 503 when ``GROQ_API_KEY`` is not configured.
    """
    global _extractor
    if _extractor is None:
        try:
            _extractor = StructuredExtractor()
        except ValueError as exc:
            raise HTTPException(
                status_code=503,
                detail=f"{exc} — extraction is unavailable until it is set",
            ) from exc
    return _extractor


def _get_keyed_extractors() -> list[StructuredExtractor]:
    """Return one extractor per configured Groq key for batch sharding.

    When ``GROQ_API_KEYS`` is set, every key gets its own single-client
    ``StructuredExtractor`` so each batch shard is bound to exactly one API
    key's rate budget. Without it, the batch falls back to the shared
    extractor (single shard).

    Raises
    ------
    HTTPException
        With status 503 when no Groq key is configured.
    """
    keys = settings.groq_api_key_list
    if not keys:
        return [_get_extractor()]
    global _extractors_by_key
    if _extractors_by_key is None:
        try:
            _extractors_by_key = [
                StructuredExtractor(api_key=key) for key in keys
            ]
        except ValueError as exc:
            raise HTTPException(
                status_code=503,
                detail=f"{exc} — extraction is unavailable until it is set",
            ) from exc
    return _extractors_by_key


def _build_sku(manufacturer: str, part_number: str) -> str:
    """Build a human-readable SKU from manufacturer + part number."""
    manufacturer = (manufacturer or "").strip() or "UNKNOWN"
    part_number = (part_number or "").strip() or "PART"
    return f"{manufacturer}-{part_number}"


def _estimate_cost(text_chars: int, attribute_count: int) -> float:
    """Rough per-request cost estimate (USD) for the LLM extraction call."""
    input_tokens = text_chars / 4.0
    output_tokens = attribute_count * 40.0
    return round((input_tokens + output_tokens) * 0.0000006, 6)


FORBIDDEN_KEYS = frozenset({
    "id",
    "choices",
    "created",
    "model",
    "object",
    "system_fingerprint",
    "usage",
    "service_tier",
    "x_groq",
    "error",
    "headers",
    "status",
    "status_code",
    "tasks",
    "attributes",
})
_FORBIDDEN_FIELD_NAMES = FORBIDDEN_KEYS


def _coerce_enrichment_attribute(
    attr: Any, default_source_url: str = ""
) -> IndustrialAttribute | None:
    """Safely convert any tuple, dict, or object representation into IndustrialAttribute."""
    if attr is None:
        return None

    if isinstance(attr, IndustrialAttribute):
        if not attr.field_name or not str(attr.field_name).strip() or str(attr.field_name).lower().strip() in _FORBIDDEN_FIELD_NAMES:
            return None
        if not attr.raw_value or not str(attr.raw_value).strip():
            return None
        return attr

    if isinstance(attr, tuple):
        if len(attr) == 0:
            return None
        field_name = str(attr[0]) if attr[0] is not None else ""
        if not field_name.strip() or field_name.lower().strip() in _FORBIDDEN_FIELD_NAMES:
            return None
        raw_value = str(attr[1]) if len(attr) > 1 and attr[1] is not None else ""
        norm_val = str(attr[2]) if len(attr) > 2 and attr[2] is not None else raw_value
        unit = str(attr[3]) if len(attr) > 3 and attr[3] is not None else None
        try:
            conf = float(attr[4]) if len(attr) > 4 and attr[4] is not None else 0.9
        except (ValueError, TypeError):
            conf = 0.9
        src = str(attr[5]) if len(attr) > 5 and attr[5] is not None else default_source_url
        if not raw_value.strip():
            return None
        return IndustrialAttribute(
            field_name=field_name.strip(),
            raw_value=raw_value.strip(),
            normalized_value=norm_val.strip() if norm_val else raw_value.strip(),
            unit=unit,
            confidence_score=max(0.0, min(1.0, conf)),
            source_url=src,
        )

    if isinstance(attr, dict):
        field_name = str(attr.get("field_name", "") or "").strip()
        if not field_name or field_name.lower() in _FORBIDDEN_FIELD_NAMES:
            return None
        raw_val = attr.get("raw_value")
        if raw_val is None:
            return None
        raw_val_str = str(raw_val).strip()
        if not raw_val_str:
            return None
        norm_val = attr.get("normalized_value", raw_val_str)
        norm_val_str = str(norm_val).strip() if norm_val is not None else raw_val_str
        unit = attr.get("unit")
        try:
            conf = float(attr.get("confidence_score", attr.get("confidence", 0.9)))
        except (ValueError, TypeError):
            conf = 0.9
        src = str(attr.get("source_url", default_source_url))
        return IndustrialAttribute(
            field_name=field_name,
            raw_value=raw_val_str,
            normalized_value=norm_val_str,
            unit=str(unit) if unit is not None else None,
            confidence_score=max(0.0, min(1.0, conf)),
            source_url=src,
        )

    if hasattr(attr, "field_name") and hasattr(attr, "raw_value"):
        fn = str(getattr(attr, "field_name", "") or "").strip()
        if not fn or fn.lower() in _FORBIDDEN_FIELD_NAMES:
            return None
        rv = str(getattr(attr, "raw_value", "") or "").strip()
        if not rv:
            return None
        nv = getattr(attr, "normalized_value", rv)
        u = getattr(attr, "unit", None)
        c = getattr(attr, "confidence_score", getattr(attr, "confidence", 0.9))
        s = getattr(attr, "source_url", default_source_url)
        return IndustrialAttribute(
            field_name=fn,
            raw_value=rv,
            normalized_value=str(nv) if nv is not None else rv,
            unit=str(u) if u is not None else None,
            confidence_score=float(c) if c is not None else 0.9,
            source_url=str(s) if s else default_source_url,
        )

    return None


async def _enrich_single_product(
    request: ProductEnrichmentRequest,
    file_bytes: bytes | None = None,
    file_name: str | None = None,
    *,
    extractor: StructuredExtractor | None = None,
) -> ProductEnrichmentResponse:
    """Core enrichment pipeline shared by `/single` and batch worker threads.

    Searches and scrapes the public web for the product, parses an uploaded
    PDF datasheet if provided, builds a combined text context from the
    sources, and extracts normalized attributes via the structured LLM.
    """
    started = time.perf_counter()

    sku_id = _build_sku(request.manufacturer_name, request.part_number)

    # 1. Check versioned enrichment cache first to avoid redundant LLM/scraping tokens (unless PDF uploaded)
    if not file_bytes:
        lookup_record = ProductRecord(
            identity=ProductIdentity(
                row_id=0,
                mfg_part_number=request.part_number or "",
                manufacturer=request.manufacturer_name or "",
                raw_description=request.raw_description or "",
                category=request.category or "General",
            ),
            raw_data=RawInputData(original_row_index=0),
        )
        cached_attrs = await asyncio.to_thread(_checkpoint_store.get_enrichment_cache, lookup_record)
        if cached_attrs:
            attributes = []
            for field_name, attr_val in cached_attrs.items():
                attributes.append(
                    IndustrialAttribute(
                        field_name=field_name,
                        raw_value=attr_val.raw_value or "",
                        normalized_value=attr_val.normalized_value or attr_val.raw_value or "",
                        unit=attr_val.unit,
                        confidence_score=attr_val.confidence,
                        source_url=attr_val.source_url or _FALLBACK_SOURCE_URL,
                    )
                )
            confidence = (
                sum(attribute.confidence_score for attribute in attributes) / len(attributes)
                if attributes
                else 0.0
            )
            elapsed_ms = (time.perf_counter() - started) * 1000.0
            return ProductEnrichmentResponse(
                sku_id=sku_id,
                category=request.category,
                enriched_attributes=attributes,
                overall_confidence=round(confidence, 4),
                processing_time_ms=round(elapsed_ms, 2),
                estimated_cost_usd=0.0,
            )

    sources: list[dict[str, str]] = []

    if file_bytes:
        try:
            # PyMuPDF parsing is CPU-bound; run it off the event loop.
            pdf_text = await asyncio.to_thread(
                _pdf_parser.extract_text_and_tables, file_bytes
            )
        except ValueError as exc:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid PDF upload: {exc}",
            ) from exc
        sources.append(
            {
                "source_url": f"{_UPLOAD_SOURCE_PREFIX}{file_name or 'datasheet.pdf'}",
                "raw_content": pdf_text,
            }
        )

    try:
        scraped = await _scraper.search_and_scrape_async(
            request.manufacturer_name, request.part_number
        )
    except Exception:
        # A failed search/scrape must not sink the whole enrichment request.
        scraped = []
    sources.extend(scraped)

    chunks = [request.raw_description or ""]
    chunks.extend(source["raw_content"] for source in sources if source["raw_content"])
    combined = "\n\n".join(chunk.strip() for chunk in chunks if chunk.strip())

    if not combined.strip():
        elapsed_ms = (time.perf_counter() - started) * 1000.0
        return ProductEnrichmentResponse(
            sku_id=sku_id,
            category=request.category,
            processing_time_ms=round(elapsed_ms, 2),
        )

    source_url = (
        sources[0]["source_url"]
        if sources
        else _FALLBACK_SOURCE_URL
    )
    # Batch shards pass their own key-bound extractor; /single uses the
    # shared one. The LLM call is blocking, so it runs off the event loop,
    # gated by the global semaphore so total LLM concurrency stays bounded
    # across all requests.
    extractor = extractor if extractor is not None else _get_extractor()
    # Reserve the request's estimated input+output tokens (input chars ≈
    # tokens, capped at the extractor's input guard, plus an output margin for
    # the attribute list) before entering the concurrency semaphore, so a TPM
    # wait never holds a concurrency slot.
    estimated_tokens = min(len(combined), _MAX_INPUT_CHARS) // 3 + 400
    await _RATE_LIMITER.acquire(estimated_tokens=estimated_tokens)
    async with _LLM_CONCURRENCY:
        raw_attributes = await asyncio.to_thread(
            extractor.extract_product_specs,
            combined,
            source_url,
            request.category,
        )

    attributes: list[IndustrialAttribute] = []
    for raw_a in (raw_attributes or []):
        coerced = _coerce_enrichment_attribute(raw_a, default_source_url=source_url)
        if coerced is not None:
            attributes.append(coerced)

    confidence = (
        sum(attribute.confidence_score for attribute in attributes) / len(attributes)
        if attributes
        else 0.0
    )
    elapsed_ms = (time.perf_counter() - started) * 1000.0

    return ProductEnrichmentResponse(
        sku_id=sku_id,
        category=request.category,
        enriched_attributes=attributes,
        overall_confidence=round(confidence, 4),
        processing_time_ms=round(elapsed_ms, 2),
        estimated_cost_usd=_estimate_cost(len(combined), len(attributes)),
    )


# ---------------------------------------------------------------------------
# Health probe
# ---------------------------------------------------------------------------


@app.get("/health")
def health() -> dict[str, str]:
    """Simple liveness probe for load balancers / uptime checks."""
    return {
        "status": "ok",
        "app": settings.app_name,
        "version": settings.app_version,
    }


# ---------------------------------------------------------------------------
# POST /api/v1/enrich/single
# ---------------------------------------------------------------------------


@app.post("/api/v1/enrich/single")
async def enrich_single(request: Request) -> ProductEnrichmentResponse:
    """Enrich a single product.

    Accepts either a JSON ``ProductEnrichmentRequest`` body or a
    ``multipart/form-data`` upload carrying the same fields plus an optional
    ``file`` (PDF datasheet).
    """
    content_type = request.headers.get("content-type", "").lower()

    if content_type.startswith("application/json"):
        try:
            payload = await request.json()
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail=f"Invalid JSON body: {exc}") from exc
        try:
            product = ProductEnrichmentRequest.model_validate(payload)
        except ValidationError as exc:
            raise HTTPException(status_code=422, detail=exc.errors()) from exc
        return await _enrich_single_product(product)

    if content_type.startswith("multipart/form-data"):
        form = await request.form()
        try:
            product = ProductEnrichmentRequest(
                manufacturer_name=form.get("manufacturer_name"),
                part_number=form.get("part_number"),
                raw_description=form.get("raw_description") or None,
                category=form.get("category"),
            )
        except ValidationError as exc:
            raise HTTPException(status_code=422, detail=exc.errors()) from exc

        upload: UploadFile | None = form.get("file")
        file_bytes = await upload.read() if upload is not None else None
        return await _enrich_single_product(
            product, file_bytes=file_bytes, file_name=upload.filename if upload else None
        )

    raise HTTPException(
        status_code=415,
        detail="Expected 'application/json' or 'multipart/form-data' content type",
    )


# ---------------------------------------------------------------------------
# POST /api/v1/enrich/batch
# ---------------------------------------------------------------------------


#: Strips punctuation from batch header keys so alternate spellings
#: ("Manufacturer Name", "manufacturer_name", "Mfr") match the same column.
_HEADER_PUNCTUATION_RE = re.compile(r"[^a-z0-9]")


def _normalize_header(key: str) -> str:
    """Lowercase *key* and strip punctuation for header-tolerant matching."""
    return _HEADER_PUNCTUATION_RE.sub("", str(key).strip().lower())


def _clean_cell(row: dict[str, str], *aliases: str) -> str:
    """Return the first non-empty cell among *aliases* (header-tolerant).

    Header keys are compared case-insensitively with punctuation stripped, so
    "Manufacturer", "Manufacturer_Name", "manufacturer name" and "mfr" all
    resolve to the same column.
    """
    lookup = {_normalize_header(key): value for key, value in row.items()}
    for alias in aliases:
        value = lookup.get(_normalize_header(alias))
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


#: Trailing code-like parentheses sometimes found on manufacturer names in raw
#: inventory dumps (e.g. "Freud Inc (2435)", "Mirka Abrasives Inc (MIRUS)")
#: are stripped so the real manufacturer name is enriched.
_PARENTHETICAL_SUFFIX_RE = re.compile(r"\s*\([^)]*\)$")


def _strip_parenthetical_suffix(text: str) -> str:
    """Strip a trailing code-like parenthetical from *text*.

    ``"Freud Inc (2435)"`` -> ``"Freud Inc"``, ``"Mirka Abrasives Inc
    (MIRUS)"`` -> ``"Mirka Abrasives Inc"``.
    """
    return _PARENTHETICAL_SUFFIX_RE.sub("", text).strip()


def _row_identity(row: dict[str, str]) -> tuple[str, str, str]:
    """Return ``(manufacturer, part_number, category)`` from a batch row.

    Manufacturer aliases are matched flexibly so any of ``manufacturer_name``,
    ``manufacturer``, ``mfr``, ``brand``, ``part_manuf``,
    ``part_manufacturer`` or ``mfg_manuf`` (in any header casing/spacing)
    supplies the real manufacturer instead of an "UNKNOWN" fallback, with
    trailing code parentheses stripped. Missing/empty categories default to
    ``"General"`` so rows without a category column still enrich.
    """
    manufacturer = _clean_cell(
        row,
        "manufacturer_name",
        "manufacturer",
        "mfr",
        "brand",
        "part_manuf",
        "part_manufacturer",
        "mfg_manuf",
    )
    manufacturer = _strip_parenthetical_suffix(manufacturer)
    return (
        manufacturer,
        _clean_cell(
            row,
            "part_number",
            "mfg_part_num",
            "part_num",
            "part_no",
            "sku",
            "part",
        ),
        _clean_cell(row, "category") or "General",
    )


def _row_to_request(row: dict[str, str]) -> ProductEnrichmentRequest:
    """Map a batch row to a request.

    Supports the official hackathon headers (``Mfg_Part_Num``/``Part_Manuf``/
    ``Part_Desc``) as well as the human-friendly aliases
    (``Manufacturer``/``Part_Number``/``Category``).
    """
    manufacturer, part_number, category = _row_identity(row)
    return ProductEnrichmentRequest(
        manufacturer_name=manufacturer,
        part_number=part_number,
        raw_description=(
            _clean_cell(row, "description", "raw_description", "part_desc", "part_description")
            or None
        ),
        category=category,
    )


def _chunk_rows(
    rows: list[dict[str, str]], n: int
) -> list[list[dict[str, str]]]:
    """Split *rows* into at most *n* contiguous shards of roughly equal size.

    ``[r0, r1, r2, r3]`` with ``n=2`` -> ``[[r0, r1], [r2, r3]]``. Returns
    ``[]`` for empty input; with more shards than rows, trailing shards are
    simply absent (each row lands in its own shard).
    """
    if n <= 0 or not rows:
        return []
    chunk_size = math.ceil(len(rows) / n)
    return [rows[i : i + chunk_size] for i in range(0, len(rows), chunk_size)]


def _read_batch_rows(filename: str, raw: bytes) -> list[dict[str, str]]:
    """Parse an uploaded CSV or Excel file into a list of header-keyed rows.

    Raises
    ------
    HTTPException
        Status 415 for unsupported extensions, 400 for empty/unreadable files.
    """
    lowered = (filename or "").lower()

    if lowered.endswith(".csv"):
        text = raw.decode("utf-8-sig", errors="replace")
        rows = list(csv.DictReader(io.StringIO(text)))
    elif lowered.endswith(".xlsx"):
        try:
            workbook = load_workbook(
                io.BytesIO(raw), read_only=True, data_only=True
            )
        except Exception as exc:
            raise HTTPException(
                status_code=400, detail=f"Could not read spreadsheet: {exc}"
            ) from exc
        try:
            sheet = workbook.active
            try:
                header = [
                    str(cell).strip() if cell is not None else ""
                    for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                ]
            except StopIteration:
                raise HTTPException(
                    status_code=400,
                    detail="Spreadsheet is empty; expected a header row",
                )
            rows = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                if all(value is None for value in values):
                    continue
                rows.append(
                    dict(zip(header, ["" if value is None else str(value) for value in values]))
                )
        finally:
            workbook.close()
    else:
        raise HTTPException(
            status_code=415,
            detail="Batch upload must be a '.csv' or '.xlsx' file",
        )

    if len(rows) > _MAX_BATCH_ROWS:
        raise HTTPException(
            status_code=422,
            detail=f"Batch upload exceeds the {_MAX_BATCH_ROWS}-row limit",
        )
    return rows


_BATCH_JOBS: dict[str, dict[str, Any]] = {}


async def _process_batch_background(
    job_id: str,
    rows: list[dict[str, str]],
    extractors: list[StructuredExtractor],
    chunk_size: int = 4,
) -> None:
    """Process batch rows in background micro-chunks (default 4) with sequential key rotation,
    cooldown wait logic, and immediate SQLite CheckpointStore persistence.
    """
    job = _BATCH_JOBS.get(job_id)
    if not job:
        return

    chunk_list = [rows[i : i + chunk_size] for i in range(0, len(rows), chunk_size)]
    extractor_cycle = extractors or [_get_extractor()]

    # Multi-key pool for cooldown checks
    groq_keys = settings.groq_api_key_list
    all_keys = groq_keys or []

    total_succeeded = 0
    total_failed = 0

    for chunk_idx, chunk_rows in enumerate(chunk_list):
        # 1. Cooldown check: if all configured keys are on cooldown, pause cleanly
        if all_keys:
            wait_seconds = _RATE_LIMITER.get_cooldown_wait(all_keys)
            if wait_seconds > 0:
                minutes = wait_seconds / 60.0
                logger.info(
                    "All API keys on TPD/TPM cooldown. Pausing processing for %.1f minutes until earliest key reopens...",
                    minutes,
                )
                await asyncio.sleep(wait_seconds)

        # 2. Sequential Key Assignment across available extractors
        extractor = extractor_cycle[chunk_idx % len(extractor_cycle)]
        # Limit concurrent row processing to maximum 2 simultaneous row executions to prevent memory spikes
        semaphore = asyncio.Semaphore(2)

        async def process_row(row: dict[str, str], row_id: int) -> BatchItemResult:
            try:
                product = _row_to_request(row)
            except Exception as exc:
                manufacturer, part_number, category = _row_identity(row)
                return BatchItemResult(
                    sku_id=_build_sku(manufacturer, part_number),
                    manufacturer_name=manufacturer,
                    part_number=part_number,
                    category=category,
                    status="error",
                    error=str(exc),
                )

            lookup_record = ProductRecord(
                identity=ProductIdentity(
                    row_id=row_id,
                    mfg_part_number=product.part_number or "",
                    manufacturer=product.manufacturer_name or "",
                    raw_description=product.raw_description or "",
                    category=product.category or "General",
                ),
                raw_data=RawInputData(original_row_index=row_id - 1),
            )

            # 1. Pre-Execution Cache Guard: Check enrichment cache BEFORE acquiring semaphore or LLM calls
            cached_attrs = await asyncio.to_thread(_checkpoint_store.get_enrichment_cache, lookup_record)
            if cached_attrs:
                attributes = [
                    IndustrialAttribute(
                        field_name=field_name,
                        raw_value=attr_val.raw_value or "",
                        normalized_value=attr_val.normalized_value or attr_val.raw_value or "",
                        unit=attr_val.unit,
                        confidence_score=attr_val.confidence,
                        source_url=attr_val.source_url or _FALLBACK_SOURCE_URL,
                    )
                    for field_name, attr_val in cached_attrs.items()
                ]
                confidence = (
                    sum(a.confidence_score for a in attributes) / len(attributes)
                    if attributes
                    else 0.0
                )
                res = BatchItemResult(
                    sku_id=_build_sku(product.manufacturer_name, product.part_number),
                    manufacturer_name=product.manufacturer_name,
                    part_number=product.part_number,
                    category=product.category,
                    status="success",
                    enriched_attributes=attributes,
                    overall_confidence=round(confidence, 4),
                    processing_time_ms=0.5,
                )
                lookup_record.attributes = cached_attrs
                lookup_record.status = RowStatus.COMPLETED
                lookup_record.enrichment_source = "enrichment_cache"
                await asyncio.to_thread(_checkpoint_store.save_checkpoint, lookup_record)
                return res

            async with semaphore:
                try:
                    response = await _enrich_single_product(
                        product, extractor=extractor
                    )
                    res = BatchItemResult(
                        sku_id=response.sku_id,
                        manufacturer_name=product.manufacturer_name,
                        part_number=product.part_number,
                        category=product.category,
                        status="success",
                        enriched_attributes=response.enriched_attributes,
                        overall_confidence=response.overall_confidence,
                        processing_time_ms=response.processing_time_ms,
                    )
                    attr_values: dict[str, AttributeValue] = {}
                    for raw_a in (res.enriched_attributes or []):
                        a = _coerce_enrichment_attribute(raw_a)
                        if a is not None and a.field_name:
                            attr_values[a.field_name] = AttributeValue(
                                field_name=a.field_name,
                                raw_value=a.raw_value,
                                normalized_value=a.normalized_value or a.raw_value,
                                unit=a.unit,
                                confidence=a.confidence_score,
                                source=ExtractionSource.LLM_8B,
                            )
                    rec = ProductRecord(
                        identity=ProductIdentity(
                            row_id=row_id,
                            mfg_part_number=product.part_number or "",
                            manufacturer=product.manufacturer_name or "",
                            raw_description=product.raw_description or "",
                            category=product.category or "General",
                        ),
                        raw_data=RawInputData(original_row_index=row_id - 1),
                        attributes=attr_values,
                        status=RowStatus.COMPLETED,
                        enrichment_source="llm_8b",
                    )
                    await asyncio.to_thread(_checkpoint_store.save_checkpoint, rec)
                    await asyncio.to_thread(_checkpoint_store.save_enrichment_cache, rec)
                    return res
                except Exception as exc:
                    manufacturer, part_number, category = _row_identity(row)
                    return BatchItemResult(
                        sku_id=_build_sku(manufacturer, part_number),
                        manufacturer_name=manufacturer,
                        part_number=part_number,
                        category=category,
                        status="error",
                        error=str(exc),
                    )

        tasks = []
        for offset, row in enumerate(chunk_rows):
            row_id = chunk_idx * chunk_size + offset + 1
            tasks.append(asyncio.create_task(process_row(row, row_id)))
            if len(tasks) < _SHARD_CONCURRENCY:
                await asyncio.sleep(_SHARD_DISPATCH_DELAY)

        chunk_results = list(await asyncio.gather(*tasks))

        # 3. Publish chunk results immediately
        for res in chunk_results:
            job["records"].append(res.model_dump())
            if res.status == "success":
                total_succeeded += 1
            else:
                total_failed += 1

        job["completed_rows"] = len(job["records"])
        job["succeeded_count"] = total_succeeded
        job["failed_count"] = total_failed

        success_confidences = [
            r.get("overall_confidence") or 0.0
            for r in job["records"]
            if r.get("status") == "success"
        ]
        if success_confidences:
            job["avg_confidence"] = round(
                sum(success_confidences) / len(success_confidences), 4
            )
        else:
            job["avg_confidence"] = 0.0

        # Persist chunk progress to SQLite
        await asyncio.to_thread(_checkpoint_store.save_batch_job, job)

        # Trigger garbage collection between micro-chunks to keep memory strictly < 250MB
        import gc
        gc.collect()

    job["is_complete"] = True
    await asyncio.to_thread(_checkpoint_store.save_batch_job, job)


@app.post("/api/v1/enrich/batch")
async def enrich_batch(
    file: UploadFile = File(...),
    sync: bool = Query(default=False),
) -> Any:
    """Enrich many products from an uploaded CSV / Excel file.

    Returns HTTP 202 Accepted with a job_id for async 4-row chunk background
    processing, pollable via `GET /api/v1/enrich/batch/{job_id}/status`.
    """
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    extractors = _get_keyed_extractors()
    rows = await asyncio.to_thread(_read_batch_rows, file.filename or "", raw)

    job_id = str(uuid.uuid4())
    job_record = {
        "job_id": job_id,
        "total_rows": len(rows),
        "completed_rows": 0,
        "succeeded_count": 0,
        "failed_count": 0,
        "is_complete": False,
        "avg_confidence": 0.0,
        "records": [],
        "created_at": time.time(),
    }
    _BATCH_JOBS[job_id] = job_record
    await asyncio.to_thread(_checkpoint_store.save_batch_job, job_record)

    if sync:
        await _process_batch_background(job_id, rows, extractors, chunk_size=4)
        job = _BATCH_JOBS[job_id]
        return BatchEnrichmentResponse(
            total=job["total_rows"],
            succeeded=job["succeeded_count"],
            failed=job["failed_count"],
            results=[BatchItemResult.model_validate(r) for r in job["records"]],
        )

    asyncio.create_task(_process_batch_background(job_id, rows, extractors, chunk_size=4))

    return {
        "job_id": job_id,
        "total_rows": len(rows),
        "status": "processing",
    }


@app.get("/api/v1/enrich/batch/{job_id}/status")
async def get_batch_status(job_id: str) -> dict[str, Any]:
    """Return current job processing state and records for a batch upload."""
    job = _BATCH_JOBS.get(job_id)
    if not job:
        job = await asyncio.to_thread(_checkpoint_store.get_batch_job, job_id)
        if job:
            _BATCH_JOBS[job_id] = job
    if not job:
        raise HTTPException(status_code=404, detail=f"Batch job '{job_id}' not found")
    return {
        "job_id": job["job_id"],
        "total_rows": job["total_rows"],
        "completed_rows": job["completed_rows"],
        "succeeded_count": job["succeeded_count"],
        "failed_count": job["failed_count"],
        "is_complete": job["is_complete"],
        "avg_confidence": job["avg_confidence"],
        "records": job["records"],
    }


# ---------------------------------------------------------------------------
# POST /api/v1/export/excel
# ---------------------------------------------------------------------------


def _style_sheet(workbook: Workbook, title: str, widths: list[int]) -> None:
    """Apply bold headers, sane column widths and frozen panes to *title*."""
    sheet = workbook[title]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
    sheet.freeze_panes = "A2"


def _build_workbook(
    products: list[tuple[ProductEnrichmentResponse, str, str]] | list[Any],
) -> Workbook:
    """Render enriched products into a flat 252-column Unilog delivery schema Excel workbook."""
    processed_items: list[dict[str, Any]] = []
    for item in products:
        if isinstance(item, tuple) and len(item) == 3:
            resp, mfg, part = item
            d = resp.model_dump()
            d["manufacturer_name"] = mfg or d.get("manufacturer_name", "")
            d["part_number"] = part or d.get("part_number", "")
            processed_items.append(d)
        elif isinstance(item, ProductEnrichmentResponse):
            processed_items.append(item.model_dump())
        else:
            processed_items.append(item)
    return CatalogExporter.build_unilog_workbook(processed_items)


def _render_workbook(
    products: list[tuple[ProductEnrichmentResponse, str, str]],
) -> io.BytesIO:
    """Build and serialize a workbook inside a single worker thread.

    ``openpyxl`` objects are not thread-safe, so building *and* saving must
    happen in the same thread — ``asyncio.to_thread`` guarantees this by
    running the whole function in one executor call.
    """
    buffer = io.BytesIO()
    workbook = _build_workbook(products)
    try:
        workbook.save(buffer)
    finally:
        workbook.close()
    buffer.seek(0)
    return buffer


@app.post("/api/v1/export/excel")
async def export_excel(
    request: Request,
    data: str | None = Query(
        default=None,
        description="JSON-encoded array of enriched product records. "
        "Alternative (preferred): send the array as the JSON request body.",
    ),
) -> StreamingResponse:
    """Render enriched results as a downloadable ``.xlsx`` workbook.

    Expects a JSON array of enriched product records, sent either as the
    request body (recommended — avoids URL-length limits on large batches)
    or as the ``data`` query parameter.
    """
    if data is not None:
        try:
            payload = json.loads(data)
        except json.JSONDecodeError as exc:
            raise HTTPException(
                status_code=400, detail=f"Invalid JSON in 'data' query parameter: {exc}"
            ) from exc
    else:
        raw_body = await request.body()
        if not raw_body:
            raise HTTPException(
                status_code=422,
                detail="Provide enriched results via a JSON request body "
                "(array of enriched product records) or the 'data' query parameter",
            )
        try:
            payload = json.loads(raw_body)
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail=f"Invalid JSON body: {exc}") from exc

    if not isinstance(payload, list):
        raise HTTPException(
            status_code=422, detail="Enriched results must be a JSON array"
        )

    products: list[tuple[ProductEnrichmentResponse, str, str]] = []
    for record in payload:
        if not isinstance(record, dict):
            raise HTTPException(
                status_code=422, detail="Every enriched product must be a JSON object"
            )
        try:
            response = ProductEnrichmentResponse.model_validate(record)
        except ValidationError as exc:
            raise HTTPException(
                status_code=422,
                detail=f"Invalid enriched product record: {exc.errors()}",
            ) from exc
        products.append(
            (
                response,
                str(record.get("manufacturer_name", "")),
                str(record.get("part_number", "")),
            )
        )

    # openpyxl rendering is CPU-bound; build + save in one worker thread.
    buffer = await asyncio.to_thread(_render_workbook, products)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": 'attachment; filename="enrichment_export.xlsx"'
        },
    )
